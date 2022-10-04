#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import bs4
import requests
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from selenium.webdriver.support import expected_conditions as EC
from datetime import date
import openpyxl
from openpyxl.styles import Font
import time

#Depurar errores de selenium en windows
options = webdriver.ChromeOptions() 
options.add_experimental_option("excludeSwitches", ["enable-logging"])

def extraer_cuil(i, sheet_obj,path):
  #del excel
  cell_obj = sheet_obj.cell(row = i, column = 2)
  return cell_obj.value

def extraer_clave(i, sheet_obj, path):
  # del excel
  cell_obj = sheet_obj.cell(row = i, column = 3)
  return cell_obj.value

def login(browser, cuit, clave_fiscal):
  try:
    browser.get('https://auth.afip.gob.ar/contribuyente_/login.xhtml')
    browser.find_element(By.ID, 'F1:username').send_keys(cuit)
    browser.find_element(By.ID, 'F1:btnSiguiente').click()
    browser.find_element(By.ID, 'F1:password').send_keys(clave_fiscal)
    browser.find_element(By.ID, 'F1:btnIngresar').click()
    print('Login exitoso.\n')
  except:
    print('Error en login.\n')

def re_login(browser, clave_fiscal):
  wait = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.ID, 'F1:btnSiguiente')))
  browser.find_element(By.ID, 'F1:btnSiguiente').click()
  browser.find_element(By.ID, 'F1:password').send_keys(clave_fiscal)
  browser.find_element(By.ID, 'F1:btnIngresar').click()    

def siper(browser):
  try:
    wait = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'p.small.light')))
    return browser.find_elements(By.CSS_SELECTOR, 'p.small.light')[2].text
  except:
    print('No se encontro elemento siper.\n')

def str_a_saldo(string):
  new_string = ''
  for s in string[2:]:
    if s.isnumeric(): new_string += s
    if s == ',': new_string += '.'
  return float(new_string)


def deuda(browser, cuit, clave_fiscal):
  try:
    # login de nuevo
    browser.get('https://ctacte.cloud.afip.gob.ar/contribuyente/externo')
    re_login(browser, clave_fiscal)
    # desplegable 
    cuits = Select(browser.find_element(By.NAME,'$PropertySelection'))
    total = {}
    for i, cuit in enumerate(cuits.options):
      cant_imp, saldo = 0, 0
      num_cuit = cuit.text
      print(f'Leyendo info de CUIT: {num_cuit}')
      total[num_cuit] = []
      time.sleep(4)
      cuits.select_by_index(i)
      time.sleep(4)
      conceptos = browser.find_elements(By.CSS_SELECTOR, 'tr[class="group"]')
      print(f'Conceptos encontrados: {len(conceptos)}')
      for concepto in conceptos:
        elemento_cant_imp = concepto.find_element(By.CSS_SELECTOR, 'span[class="cant-impuesto"]').text
        cant_imp += int(elemento_cant_imp[1:-1])
        elemento_saldo = concepto.find_element(By.CSS_SELECTOR, 'td[class="subtotales sb_saldo"]').text
        saldo += str_a_saldo(elemento_saldo)
        elemento_int_r = concepto.find_element(By.CSS_SELECTOR, 'td[class="subtotales sb_int_res"]').text
        saldo += str_a_saldo(elemento_int_r)
        elemento_int_p = concepto.find_element(By.CSS_SELECTOR, 'td[class="subtotales sb_int_pun"]').text
        saldo += str_a_saldo(elemento_int_p)
      print(f'Cantidad de deuda: {saldo}')   
      print(f'Cantidad de obligaciones encontradas: {cant_imp}')
       
      total.update({num_cuit: [cant_imp, saldo]})
    return total
  except:
    print('No se encontro elemento deuda.\n')    

def e_servicios_juridicos(browser):
  try:
    browser.get('https://eservicios.srt.gob.ar/home/Default.aspx')
    return browser.find_element(By.CSS_SELECTOR, 'span.label.label-danger.badge-mensajes').text
  except:
    print("No se encontraron notificaciones juridicas. \n")

def e_servicios_personal(browser):
  try:
    browser.get('https://eservicios.srt.gob.ar/home/Default.aspx')
    re_login(browser, clave_fiscal)
    return browser.find_element(By.CSS_SELECTOR, 'span.badge.badge-light.text-right.cuit-tooltip').text
  except:
    print("No se encontraron notificaciones personales. \n")

def flattenlist(l):
  flat_list = []
  for sublist in l:
    for item in sublist:
        flat_list.append(item)
  return flat_list          

if __name__ == '__main__':
  #extraccion de datos del excel
  path = r"C:\Users\Administrator\Desktop\AFIP\afip\AFIP.xlsx"
  wb_obj = openpyxl.load_workbook(path)  
  sheet_obj = wb_obj.active
  max_row = sheet_obj.max_row
  #extrae cada usuario del excel
  for i in range(2, max_row - 1):
    cuit = extraer_cuil(i, sheet_obj, path)
    clave_fiscal = extraer_clave(i, sheet_obj, path)
    browser = webdriver.Chrome(options=options)
    login(browser, cuit, clave_fiscal)
    #A partir de aca solo realizar si el login fue exitoso
    riesgo = siper(browser)
    #Escribir siper en excel
    c1 = sheet_obj.cell(row = i, column = 4)
    c1.value = riesgo
    # deuda
    deuda_dic = deuda(browser, cuit, clave_fiscal)
    if type(deuda_dic) == dict:
      deuda_list = list(deuda_dic.values())
      deuda_list = flattenlist(deuda_list)
      for j in range(len(deuda_list)):
        c1 = sheet_obj.cell(row = i, column = 7+j)
        c1.value = deuda_list[j]
    else: 
      print("Error al encontrar elemento deuda")    
    # e-Servicios SRT
    notificaciones_personales = e_servicios_personal(browser)
    c1 = sheet_obj.cell(row = i, column = 5)
    c1.value = notificaciones_personales
    notificaciones_juridicas = e_servicios_juridicos(browser)
    c1 = sheet_obj.cell(row = i, column = 6)
    c1.value = notificaciones_juridicas
    wb_obj.save(path)
  browser.close()  